from pathlib import Path
from openpyxl import load_workbook


downloads = Path.home() / "Downloads"
xlsx_files = [p for p in downloads.glob("*.xlsx") if "26.4.13" in p.name]
if not xlsx_files:
    raise SystemExit("No matching xlsx found")

p = xlsx_files[0]
print(f"XLSX: {p}")
wb = load_workbook(p, data_only=True)
print("SHEETS:", wb.sheetnames)
for ws in wb.worksheets:
    print(f"\nSHEET: {ws.title} rows={ws.max_row} cols={ws.max_column}")
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 25), values_only=True):
        print(row)
